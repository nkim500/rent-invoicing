import base64
import os
import re
import shutil
import zipfile
from datetime import date
from datetime import datetime
from datetime import timedelta
from io import BytesIO
from uuid import UUID

import data_models as models
import numpy as np
import pandas as pd
import streamlit as st
from config import BusinessEntityParams
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from pydantic import ValidationError


def statement_date_widget() -> date:
    """
    Sets the statement date for the session that the user is working on. The function
    creates Streamlit number input widgets to take user input for a year and month to
    return the first of the selected month as date object.
    """
    st.sidebar.write("Select the statement date to work on:")
    year = st.sidebar.number_input(
        label="Year", min_value=2020, max_value=2100, value=date.today().year
    )
    month = st.sidebar.number_input(
        label="Month", min_value=1, max_value=12, value=date.today().month
    )
    selected_date = date(year, month, 1)

    if st.session_state.statement_date != selected_date:
        st.session_state.statement_date = selected_date
        st.session_state.invoice_setting_index = None

    return selected_date


def format_func_invoice_settings(invoice_setting: dict):
    """Formatting function for displaying the available invoice setting options"""
    return f"""
        Date: {
            invoice_setting["inserted_at"][:10]} - Rent ${invoice_setting["rent_monthly_rate"]:.2f} / Storage ${invoice_setting["storage_monthly_rate"]:.2f} / Water_rate ${invoice_setting["water_monthly_rate"]:.6f} / Water_fee ${invoice_setting["water_service_fee"]:.2f} / Late_fee {invoice_setting["late_fee_rate"]*100:.1f}%
    """  # noqa: E501


def invoice_setting_widget(invoice_settings: list[dict]) -> dict:
    """
        Sets the session invoice setting to the first selection with an 'effective as of'
        date earlier than the session statement date.

        User can override the automatically selected session invoice setting by manually
        selecting via the Streamlit select box, which this function also creates.

    Args:
        invoice_settings (list[dict]):
            List of invoice settings available in the database. By system default, the
            input list is likely sorted by database insertion date, in descending order.

    Returns:
        dict: Dict form of an invoice setting object
    """
    if st.session_state.statement_date and st.session_state.invoice_setting_index is None:
        matching_index = next(
            (
                i
                for i, setting in enumerate(invoice_settings)
                if (
                    date.fromisoformat(setting["effective_as_of"])
                    <= st.session_state.statement_date
                )
            ),
            None,
        )
        st.session_state.invoice_setting_index = (
            matching_index if matching_index is not None else None
        )

    st.session_state.invoice_setting = st.selectbox(
        "Select the invoice configuration",
        options=invoice_settings,
        format_func=format_func_invoice_settings,
        index=st.session_state.invoice_setting_index,
    )

    return st.session_state.invoice_setting


def properties_widget(properties: list[dict]) -> dict:
    """Returns a widget that allows user to set the session property to work with

    Args:
        properties (list[dict]): List of properties available in the database

    Returns:
        dict: Dict form of a property object
    """
    st.session_state.prop = st.sidebar.selectbox(
        "Select the invoice configuration",
        options=properties,
        format_func=lambda x: x["property_code"],
    )

    return st.session_state.prop


def extract_lot_number(lot_id: str | None) -> int | None:
    """Formatting function for composing invoices"""
    if lot_id:
        return int(re.sub(r"\D", "", lot_id))
    else:
        return lot_id


def preview_payments_dataframe(
    payments: list[dict], acccounts_with_tenant_info: list[dict] | None = None
) -> pd.DataFrame:
    """Returns a pandas.DataFrame of payments

    Args:
        payments (list[dict]):
            List of [models.Payment.model_dump()]
        acccounts_with_tenant_info (list[dict] | None, optional):
            If the API request providing the argument had with_tenant_info set to True,
            indicate True. Defaults to None.
    """
    df = pd.DataFrame(list(payments))
    df["amount_available"] = df["amount"] - df["amount_applied"]
    pmt_columns = [
        "id",
        "inserted_at",
        "payer",
        "beneficiary_account_id",
        "amount_available",
        "amount",
        "payment_received",
    ]
    df = df[pmt_columns]

    if acccounts_with_tenant_info:
        acct = pd.DataFrame(list(acccounts_with_tenant_info))
        acct.rename(
            columns={"id": "account_id", "full_name": "Payment for"}, inplace=True
        )
        df = df.merge(
            acct, how="left", left_on="beneficiary_account_id", right_on="account_id"
        )
        df = df[pmt_columns[:3] + ["Payment for", "lot_id"] + pmt_columns[3:]]
        df = df.drop(columns=["beneficiary_account_id"])

    df.rename(
        columns={
            "inserted_at": "Inserted at",
            "payment_received": "Received on",
            "payer": "Payment made by",
            "amount": "Original amount",
            "amount_available": "Available amount",
        },
        inplace=True,
    )

    return df


def preview_charges_dataframe(
    receivables: dict,
    accounts: list[dict],
    payments: list[dict] | None = None,
) -> pd.DataFrame:
    """
        Returns a pandas.DataFrame of current balance, existing or new charges, payments
        available, resulting late fees, and resulting balance.

    Args:
        receivables (dict): mapping of account_ids to receivables
        accounts (list[dict]): list of accounts
        payments (list[dict] | None, optional): available payments. Defaults to None.

    Returns:
        pd.DataFrame: _description_
    """
    df_new = pd.DataFrame.from_dict(
        {
            k: {
                item["charge_type"]: sum(
                    i["amount_due"]
                    for i in v[0]
                    if i["charge_type"] == item["charge_type"]
                )
                for item in v[0]
                if item is not None
            }
            for k, v in receivables.items()
        },
        orient="index",
    )
    df_new = df_new[sorted(df_new.columns, reverse=True)]
    df_new["total_new"] = df_new.sum(axis=1)
    df_in_db = pd.DataFrame.from_dict(
        {
            k: sum(item["amount_due"] for item in v[1] if item["paid"] is False)
            if v[1]
            else None
            for k, v in receivables.items()
        },
        orient="index",
        columns=["current_outstanding"],
    )
    output_df = df_in_db.merge(df_new, how="left", left_index=True, right_index=True)
    if "late_fee" in output_df.columns:
        new_cols = [col for col in output_df.columns if col != "late_fee"]
        new_cols.append("late_fee")
        output_df = output_df[new_cols]
    base_col = ["current_outstanding"]
    for col in output_df.columns:
        if col != "current_outstanding":
            base_col.append(col)
    output_df = output_df[base_col]

    df = pd.DataFrame(accounts).merge(
        output_df, how="left", left_on="id", right_index=True
    )

    label_columns = ["id", "lot_id", "full_name"]
    value_columns = [i for i in df.columns if i not in label_columns]
    df["total_due"] = df[value_columns].sum(axis=1) - df["total_new"]

    if payments:
        payments_df = pd.DataFrame(payments)
        payments_df["payment_available"] = (
            payments_df["amount"] - payments_df["amount_applied"]
        )
        payments_df.rename(columns={"beneficiary_account_id": "account_id"}, inplace=True)
        agg = payments_df.groupby("account_id").payment_available.sum()
        df = df.merge(pd.DataFrame(agg), how="left", left_on="id", right_index=True)
        df.payment_available.fillna(value=0, inplace=True)
        df["balance_after_payment"] = df.apply(
            lambda x: max(x.total_due - x.payment_available, 0), axis=1
        )

    df["lot_numbers"] = df.lot_id.apply(lambda x: extract_lot_number(x))
    df = df.sort_values(by="lot_numbers").set_index("lot_numbers")

    preview_columns = [i for i in df.columns if i != "id"]

    return df[preview_columns]


def duplicate_payment_entry_check(
    new_payments: list[models.Payment], existing_payments: list[dict]
) -> bool:
    """
        Returns True if input Payment objects for database insertion, is suspected to be
        already on the database

    Args:
        new_payments (list[models.Payment]): user created Payment object
        existing_payments (list[dict]): payments in the database

    Returns:
        bool: True, if duplicate found. Else, False.
    """
    existing_payment_objs = [
        models.Payment(**payment_dict) for payment_dict in existing_payments
    ]

    for new_payment in new_payments:
        for existing_payment in existing_payment_objs:
            if (
                new_payment.beneficiary_account_id
                == existing_payment.beneficiary_account_id
                and new_payment.amount == existing_payment.amount
                and new_payment.payment_dated == existing_payment.payment_dated
                and new_payment.payment_received == existing_payment.payment_received
                and new_payment.payer == existing_payment.payer
            ):
                return True  # Duplicate found
    return False


def duplicate_accounts_receivable_entry_check(
    new_receivables: list[models.AccountsReceivable],
    existing_receivables: list[models.AccountsReceivable],
) -> bool:
    """
        Returns False if input receivable objects for database insertion, is suspected to
        be already on the database

    Args:
        new_receivables (list[models.Payment]): user created receivable object
        existing_receivables (list[dict]): receivables in the database

    Returns:
        bool: True, if duplicate found. Else, False.
    """
    for new_receivable in new_receivables:
        for existing_receivable in existing_receivables:
            if (
                new_receivable.account_id == existing_receivable.account_id
                and new_receivable.amount_due == existing_receivable.amount_due
                and new_receivable.statement_date == existing_receivable.statement_date
                and new_receivable.charge_type.value
                == existing_receivable.charge_type.value
            ):
                return True  # Duplicate found
    return False


def remove_empty_rows(ws: Worksheet) -> Worksheet:
    """
    Takes a draft invoice in Worksheet format and returns the same with the empty rows
    in the account activity section.
    """
    index_row = []
    add_back_count = 0
    add_back_start_row = 31

    for i in range(13, 21):
        if ws.cell(i, 3).value is None:
            index_row.append(i)
            add_back_count += 1
            add_back_start_row -= 1

    for row_del in range(len(index_row)):
        ws.delete_rows(idx=index_row[row_del], amount=1)
        index_row = [k - 1 for k in index_row]

    ws.insert_rows(add_back_start_row, add_back_count)

    return ws


def generate_invoices(
    template_path: str,
    input_data: list[models.InvoiceFileParse],
    export_path: str,
):
    """Generates and saves invoices locally

    Args:
        template_path (str): local directory containing the template invoice file
        input_data (list[models.InvoiceFileParse]): invoice data to populate the invoices
        export_path (str): local directory to save the composed invoice files

    Returns:
        _type_: _description_
    """
    export_file_paths = []
    for i in input_data:
        wb = load_workbook(template_path)
        ws = wb.active
        model = i.model_dump()
        for k, v in model.items():
            ws[k] = v
        ws = remove_empty_rows(ws)
        export_file_path = f"{export_path}{i.F4} Bill {i.F6.strftime("%b %Y")}.xlsx"
        export_file_paths.append(export_file_path)
        wb.save(export_file_path)
        wb.close()
    return export_file_paths


def ingest_water_meter_readings(report_file: BytesIO) -> pd.DataFrame:
    """Ingests water usage report in .xlsx format from user and returns as pd.DataFrame

    Args:
        report_file (BytesIO): uploaded water meter report
    """
    df = pd.read_excel(report_file, header=1, index_col=0)
    try:
        previous_date = df.columns[3]
        current_date = df.columns[2]
        assert isinstance(previous_date, date)
        assert isinstance(current_date, date)
        if isinstance(previous_date, datetime) and isinstance(current_date, datetime):
            df.rename(
                columns={
                    df.columns[3]: df.columns[3].date(),
                    df.columns[2]: df.columns[2].date(),
                },
                inplace=True,
            )
    except AssertionError as e:
        raise (e)

    return df


def generate_water_usage_objects(
    report: pd.DataFrame, statement_date: date | None = None, indexed: bool = False
) -> list[models.WaterUsage] | dict[int, models.WaterUsage]:
    """Composes a list of WaterUsage objects from a pd.DataFrame of water usage report

    Args:
        report (pd.DataFrame):
            Water usage report
        statement_date (date | None, optional):
            The statement date to assign to the composed water usage objects. If None, the
            statement date will set to the first day of current month. Defaults to None.

    Returns:
        list[models.WaterUsage]: composed WaterUsage objects
    """
    if not statement_date:
        statement_date = date.today().replace(day=1)

    water_usages = []
    if indexed:
        water_usages = {}

    check_list = []

    previous_date = report.columns[3]
    current_date = report.columns[2]

    for n, row in report.iterrows():
        try:
            water_usage = models.WaterUsage(
                watermeter_id=row["Meter #"],
                previous_date=previous_date,
                current_date=current_date,
                statement_date=statement_date,
                previous_reading=row.iloc[3],
                current_reading=row.iloc[2],
            )
            if indexed:
                water_usages[n] = water_usage
            else:
                water_usages.append(water_usage)
        except ValidationError:
            check_list.append(row.name)

    if check_list:
        return check_list

    return water_usages


def ingest_bookkeeping_excel(
    book: BytesIO, sheet_name: str | None = None
) -> pd.DataFrame:
    """Ingests bookkeeping file in .xlsx format from user and returns as pd.DataFrame

    Args:
        report_file (BytesIO): uploaded bookkeeping file
    """
    if sheet_name:
        df = pd.read_excel(book, header=2, index_col=0, sheet_name=sheet_name)
    else:
        df = pd.read_excel(book, header=2, index_col=0, sheet_name=-1)

    drop_col_idx = [1, 9, 15, 23, 25]
    out_df = df.iloc[:, [i for i in range(len(df.columns)) if i not in drop_col_idx]]
    out_df.columns = [
        "tenant_name",
        "starting_balance",
        "monthly_due_last_month",
        "paid_on_time_last_month",
        "paid_past_due_last_month",
        "late_fee_accrued_last_month",
        "total_carried_over_last_month",
        "ending_balance",
        "monthly_rent",
        "monthly_storage",
        "monthly_water",
        "monthly_other",
        "new_charges_this_month",
        "payment_on_time_1",
        "payment_on_time_2",
        "payment_on_time_3",
        "payment_overdue_1",
        "payment_overdue_2",
        "payment_overdue_3",
        "payment_overdue_4",
        "late_fee_this_month",
        "carry_over_to_next_month",
    ]

    return out_df


def serialize_invoice_input_from_book_ingest(
    company: BusinessEntityParams,
    statement_date: date,
    property_code: str,
    street_address_base: str,
    csz_address: str,
    entry: models.BookIngest,
    water: models.WaterUsage,
    invoice_setting_id: str | UUID,
    # as_invoice_object: bool = False,
) -> dict:
    """
    Parse from bookkeeping .xlsx files and convert it into a dictionary or Invoice object.

    Args:
        book_entries (list[BookIngest]):
            List containing invoice data fields as BookIngest object.
        as_invoice_object (bool, optional):
            If True, returns a validated Invoice model object instead of a dictionary.

    Returns:
        dict | list[models.Invoice]:
            Parsed row data in dictionary form or as an Invoice object.
    """
    statement_date = statement_date
    total_amount_due = entry.total_amount_due_for_invoice
    invoice_setting_id = invoice_setting_id

    if not total_amount_due:
        return

    lot_id = property_code + str(entry.lot_id)
    csz = csz_address
    if lot_id:
        tenant_address_1 = str(entry.lot_id) + " " + street_address_base
    else:
        tenant_address_1 = ""
    if csz:
        tenant_address_2 = csz
    else:
        tenant_address_2 = ""

    if entry.total_carried_over_last_month >= 0:
        amt_overdue_prev_mth_w_cred = entry.starting_balance
    else:
        amt_overdue_prev_mth_w_cred = (
            entry.starting_balance + entry.total_carried_over_last_month
        )

    # if entry.late_fee_accrued_last_month > 0:
    #     amt_overdue_prev_mth_w_cred = (
    #         amt_overdue_prev_mth_w_cred - entry.late_fee_accrued_last_month
    #     )

    prev_mth_total_amount_paid = (
        entry.paid_on_time_last_month + entry.paid_past_due_last_month
    )

    parsed = {
        "invoice_customer_id": f"{lot_id}",
        "tenant_address_1": tenant_address_1,
        "tenant_address_2": tenant_address_2,
        "tenant_name": entry.tenant_name,
        "amt_prev_month_paid": prev_mth_total_amount_paid,
        "amt_prev_month_residual": max(
            entry.monthly_due_last_month - prev_mth_total_amount_paid, 0
        ),
        "invoice_total_amount_due": entry.total_amount_due_for_invoice,
        "amt_total_amount_due": entry.total_amount_due_for_invoice,
        # "amt_overdue": max(entry.starting_balance, 0),
        "amt_overdue": amt_overdue_prev_mth_w_cred,
        "amt_other_rent": entry.monthly_other,
        "amt_rent": entry.monthly_rent,
        "amt_storage": entry.monthly_storage,
        "amt_water": entry.monthly_water,
        "water_bill_period": entry.monthly_water,
        "water_prev_read": water.previous_reading,
        "water_curr_read": water.current_reading,
        "water_curr_date": water.current_date,
        "water_prev_date": water.previous_date,
        "water_meter_id": water.watermeter_id,
        "amt_late_fee": entry.late_fee_accrued_last_month,
    }

    if (water.current_reading is not None) and (water.previous_reading is not None):
        parsed["desc_curr_water"] = (
            f"""Water bill for {
                parsed['water_prev_date'].strftime("%B")
            }-"""
            + f"""{parsed['water_curr_date'].strftime("%B %Y")}"""
        )
        parsed["date_water"] = statement_date
        parsed["water_usage_period"] = water.current_reading - water.previous_reading
    else:
        parsed["desc_curr_water"] = None
        parsed["date_water"] = None
        parsed["water_usage_period"] = None

    if not np.isnan(entry.late_fee_accrued_last_month) and isinstance(
        entry.late_fee_accrued_last_month, float
    ):
        parsed["desc_late_fee"] = "Late fee"
        parsed["date_late"] = statement_date
    else:
        parsed["desc_late_fee"] = None
        parsed["date_late"] = None

    if not np.isnan(entry.monthly_rent) and isinstance(entry.monthly_rent, float):
        parsed["desc_curr_rent"] = f"Lot rent for {(statement_date).strftime("%B %Y")}"
        parsed["date_rent"] = statement_date
    else:
        parsed["desc_curr_rent"] = None
        parsed["date_rent"] = None

    if not np.isnan(entry.monthly_storage) and isinstance(entry.monthly_storage, float):
        parsed["desc_curr_storage"] = f"""Storage rent for {
            (statement_date).strftime("%B %Y")
        }"""
        parsed["date_storage"] = statement_date
    else:
        parsed["desc_curr_storage"] = None
        parsed["date_storage"] = None

    total_paid_last_month = entry.paid_on_time_last_month + entry.paid_past_due_last_month

    if not np.isnan(total_paid_last_month) and isinstance(total_paid_last_month, float):
        parsed["desc_prev_month_paid"] = f"""Bill paid for {
            (statement_date - timedelta(days=28)).strftime("%B %Y")
        }"""
        parsed["date_today_1"] = models.et_date_now()
    else:
        parsed["desc_prev_month_paid"] = None
        parsed["date_today_1"] = None

    # if not np.isnan(entry.starting_balance) and entry.starting_balance > 0:
    #     parsed["desc_prev_overdue"] = "Previous overdue"
    # else:
    #     parsed["desc_prev_overdue"] = None

    # if (
    #     not np.isnan(entry.total_carried_over_last_month)
    #     and entry.total_carried_over_last_month > 0
    # ):
    #     parsed["desc_prev_month_residual"] = f"""{
    #         (statement_date - timedelta(days=28)).strftime("%B")
    #     } bill, less paid"""
    #     parsed["date_today_2"] = models.et_date_now()
    # else:
    #     parsed["desc_prev_month_residual"] = None
    #     parsed["date_today_2"] = None

    parsed["desc_prev_month_residual"] = f"""{
        (statement_date - timedelta(days=28)).strftime("%B")
    } bill, less paid"""
    parsed["date_today_2"] = models.et_date_now()

    if amt_overdue_prev_mth_w_cred:
        parsed["desc_prev_overdue"] = "Previous overdue (credit)"
    else:
        parsed["desc_prev_overdue"] = None

    if not np.isnan(entry.monthly_other) and isinstance(entry.monthly_other, float):
        parsed["desc_other_rent"] = "Other rent(s)*"
        parsed["date_other_rent"] = statement_date
        parsed["detail_other_rent"] = "* Please contact to find out the details"
    else:
        parsed["desc_other_rent"] = None
        parsed["date_other_rent"] = None
        parsed["detail_other_rent"] = None

    parsed["invoice_date"] = models.et_date_now()
    parsed["business_name"] = company.business_name
    parsed["business_address_1"] = company.business_address_1
    parsed["business_address_2"] = company.business_address_2
    parsed["business_contact_phone"] = company.business_contact_phone
    parsed["business_contact_email"] = company.business_contact_email
    parsed["invoice_due_date"] = statement_date
    parsed["business_name_"] = company.business_name.upper()
    parsed["business_address_1_"] = company.business_address_1
    parsed["business_address_2_"] = company.business_address_2
    parsed["business_contact_email_"] = f"Or Zelle to {company.business_contact_email}"
    parsed["invoice_date_"] = parsed["invoice_date"]
    parsed["invoice_customer_id_"] = parsed["invoice_customer_id"]
    parsed["invoice_due_date_"] = parsed["invoice_due_date"]
    parsed["invoice_total_amount_due_"] = parsed["invoice_total_amount_due"]

    # if as_invoice_object:
    #     return models.Invoice.model_validate(
    #         {
    #             "invoice_date": parsed["invoice_date"],
    #             "statement_date": statement_date,
    #             "account_id": row[1],
    #             "lot_id": lot_id,
    #             "tenant_name": parsed["tenant_name"],
    #             "setting_id": invoice_setting_id,
    #             "amount_due": parsed["amt_total_amount_due"],
    #             "details": parsed,
    #         }
    #     )

    return parsed


def generate_invoice_from_user_inputs(
    book: BytesIO,
    waters: BytesIO,
    statement_date: date,
    prop: dict,
    invoice_setting_id: str | UUID,
    template_path: str,
    export_path: str,
    sheet_name: str | None = None,
    company: BusinessEntityParams | None = None,
):
    if company is None:
        company = BusinessEntityParams()

    prop = models.Property.model_validate(prop)
    water_df = ingest_water_meter_readings(waters)
    water_usages = generate_water_usage_objects(water_df, statement_date, indexed=True)

    book_df = ingest_bookkeeping_excel(book=book, sheet_name=sheet_name)
    book_dict = book_df.to_dict(orient="index")
    input_data = [
        serialize_invoice_input_from_book_ingest(
            company=company,
            statement_date=statement_date,
            property_code=prop.property_code,
            street_address_base=prop.street_address,
            csz_address=prop.city_state_zip,
            entry=models.BookIngest(lot_id=lot_number, **entry),
            water=water_usages[lot_number],
            invoice_setting_id=invoice_setting_id,
        )
        for lot_number, entry in book_dict.items()
    ]
    invoice_parsed = [models.InvoiceFileParse(**i) for i in input_data if i]

    return generate_invoices(
        template_path=template_path, input_data=invoice_parsed, export_path=export_path
    )


def get_binary_file_downloader_html(bin_file, file_label="File"):
    with open(bin_file, "rb") as f:
        data = f.read()
    bin_str = base64.b64encode(data).decode()
    href = f'<a href="data:application/octet-stream;base64,{bin_str}" download="{os.path.basename(bin_file)}">Download {file_label}</a>'  # noqa: E501
    return href


def user_download_invoice_zip(file_dir: str):
    """Creates a Streamlit download button allowing user to download the composed invoices

    Args:
        file_dir (str): local (container) directory containing the invoices
    """
    zip_buffer = BytesIO()

    with zipfile.ZipFile(zip_buffer, "w") as zip_file:
        for filename in os.listdir(file_dir):
            if filename.endswith(".xlsx"):
                filepath = os.path.join(file_dir, filename)
                zip_file.write(filepath, arcname=filename)
    zip_buffer.seek(0)
    st.download_button(
        label="Download All Reports",
        data=zip_buffer,
        file_name="all_reports.zip",
        mime="application/zip",
    )
    zip_buffer.close()


def clear_directory(file_dir: str):
    """Deletes all files in the file_dir directory

    Args:
        file_dir (str): local (container) directory with the previously composed invoices
    """
    for filename in os.listdir(file_dir):
        file_path = os.path.join(file_dir, filename)
        try:
            if os.path.isfile(file_path) or os.path.islink(file_path):
                os.unlink(file_path)
            elif os.path.isdir(file_path):
                shutil.rmtree(file_path)
        except Exception as e:
            print(f"Failed to delete {file_path}. Reason: {e}")


def display_existing_invoice(invoice_data: list[dict]) -> pd.DataFrame:
    collection = []
    for invoice in invoice_data:
        details = invoice["details"]
        collection.append(
            {
                "lot": invoice["lot_id"],
                "account holder name": invoice["tenant_name"],
                "rent": details["amt_rent"],
                "storage": details["amt_storage"],
                "water": details["amt_water"],
                "other rent": details["amt_other_rent"],
                "overdue": details["amt_overdue"],
                "late fees": details["amt_late_fee"],
                "total invoice due": invoice["amount_due"],
            }
        )
    return round(pd.DataFrame(collection), 2)
